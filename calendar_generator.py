import os, requests, csv, sys
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# Set up lists to select venues and build information tables
print("Setting up the lists...")
print()
venues = ["EMO'S", "SCOOT INN"]
output = []
table = []
date_events = {}

# Compile all the data
def compile_info():
    print()
    print("Compiling info for", venue.title()+ ":")
    # Get the date of the event
    current_month = datetime.now().month
    event_months = soup.find_all('span', class_='eventMonth')
    event_days = soup.find_all('span', class_='eventDay')
    combined_dates = []

    # Establish the year and build the date variable
    for event_month, event_day in zip(event_months, event_days):
        if datetime.strptime(event_month.text, "%b").month < current_month:
            year += 1
        else:
            year = datetime.now().year
        month_numerical = datetime.strptime(event_month.text, "%b").month
        date = f"{year}-{month_numerical:02d}-{event_day.text}T00:00:00.000Z"
        combined_dates.append(date)
        year = datetime.now().year
    # Set up variables for event date list, door time list, datetime for now, next month and the month after
    combined_dates = sorted(list(set(combined_dates)), key=lambda x: datetime.strptime(x, "%Y-%m-%dT%H:%M:%S.000Z"))
    doors = soup.find_all('span', itemprop='doorTime')
    print(doors)
    now = datetime.now()
    first_next_month = (now + timedelta(days=30)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    first_after_next = (now + timedelta(days=60)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Set up the output for the CSV files
    i = 0
    for artist in artists:
        description = f"=CONCATENATE(H{len(output) + 2},UPPER(I{len(output) + 2}))"
        start_date = datetime.strptime(combined_dates[i], "%Y-%m-%dT%H:%M:%S.%fZ").date()
        # Isolate the shows for only next month
        if start_date >= first_next_month.date() and start_date < first_after_next.date():
            if i >= 0 and i < len(doors) and i < len(artists):
                combined = []
                # If the show has MOVED TO in the event title, indicating the date and/or venue location has changed
                if "MOVED TO" in artist.get_text():
                    doors.insert(i, "N/A")
                    doors_time = "N/A"
                    call_time = "N/A"
                    end_time = "N/A"
                    end_date = "N/A"
                # Otherwise, build the main lists for the remaining shows
                else:
                    doors_time = datetime.strptime(doors[i].get_text()[15:].lstrip(), '%I:%M %p')
                    # print(doors_time)
                    call_time = (doors_time - timedelta(hours=1))
                    end_dt = datetime.combine(start_date, doors_time.time()).strftime('%Y-%m-%d %I:%M %p')
                    end_datetime = (datetime.strptime(end_dt, '%Y-%m-%d %I:%M %p') + timedelta(hours=5)).strftime('%Y-%m-%d %I:%M %p')
                    end_time = end_datetime[end_datetime.index(" "):].lstrip()
                    end_date = end_datetime[0:end_datetime.index(" ")]
                # Start building the event tables that will be used to make the schedule and email the team
                event = [start_date, artist.get_text().replace("â€™", "'"), call_time.time().strftime("%I:%M %p"), venue]
                combined.extend([start_date, artist.get_text(), call_time.time().strftime("%I:%M %p"), end_time, end_date, location, description])
                combined.append(("Doors: " + doors_time.time().strftime("%I:%M %p") + "\nVenue: " + venue + "\nMerch Coordinator: ").upper())
                output.append(combined)
            else:
                # Handle the case when the index i is out of range for doors or artists
                print("\n!!! INDEX OUT OF RANGE FOR DOORS OR ARTISTS:", i, "!!!\n")
                print("What does this mean? This means that there most likely may be some data missing so you'll want to check over your csv files and first make sure that all your events are listed. Next, make sure all the event dates and doors time are correct to the event as listed on the venues' event calendar webpages. Add any events that may have gotten skipped and adjust any door times as needed.\n")
            # Build table for event date comparison
            if start_date not in date_events:
                date_events[start_date] = []
            date_events[start_date].append(event)
            print("The following show needs to be staffed:", event)
        i += 1

# Set the venues to process for the script
print("Setting up the venues for this script. The venues for this script are "  + ", ".join([venue.title() for venue in venues]) + ".")
for venue in venues:
    if venue == "EMO'S":
        url = requests.get("https://www.emosaustin.com/events-calendar")
        location = "2015 E Riverside Dr, Austin, TX 78741"
    elif venue == "SCOOT INN":
        url = requests.get("https://scootinnaustin.com/calendar")
        location = "1308 E 4th St, Austin, TX 78702"
    # You should never see the following!
    else:
        print("No more venues! Check your venue list because you shouldn't see this!")

    # Get all the HTML data from the Emo's and Scoot Inn event web pages and execute compile_info()
    soup = BeautifulSoup(url.content, 'html.parser')

    # Uncomment the following lines to print the soup variable to txt files named after the corresponding 
    # venue and see exactly what you are parsing
    # with open(f'{venue}.txt', 'w', encoding='utf-8') as python_output:
    #     # Redirect the standard output to the file
    #     sys.stdout = python_output
    #     print(soup)
    # # Restore the default standard output
    # sys.stdout = sys.__stdout__

    artists = soup.find_all('h2', itemprop='name')
    compile_info()

# Identify dates when there are shows at both venues
print()
print("Identifying dates where both venues have a show...")
print()
for date, events in date_events.items():
    if len(events) > 1:
        for event in events:
            if event[3] == "EMO'S":
                event.append("CONCURRENT SHOW AT SCOOT INN")
            elif event[3] == "SCOOT INN":
                event.append("CONCURRENT SHOW AT EMO'S")
            # You should never see the following!
            else:
                print("What d'ya want from me?")
    table.extend(events)

# Set up headers for the CSV file used to create the schedule and the table emailed to the team
print("Setting up the CSV headers...")
print()
header = ['Start Date', 'Subject', 'Start Time', 'End Time', 'End Date', 'Location', 'Description',  'Description_1', 'Merch Coordinator']
email_header = ['Start Date', 'Subject', 'Start Time', 'Venue', 'Notes']

# Generate email_table.csv in the ./csv_files folder that will be used to email the Merch Coordinators
print("Creating csv_files/email_table.csv")
with open('csv_files/email_table.csv', 'w', encoding='UTF8', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(email_header)
    for show in table:
        if any(field for field in show):
            writer.writerow(show)
print("Exported to " + os.getcwd() + "csv_files/email_table.csv")
print()

# Generate calendar.csv in the ./csv_files folder that wil be used to create the schedule in a format that can be imported in Google Calendars
print("Creating csv_files/calendar.csv")
with open('csv_files/calendar.csv', 'w', encoding='UTF8', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(header)
    for combo in output:
        if any(field for field in combo):
            writer.writerow(combo)
print("Exported to " + os.getcwd() + "csv_files/calendar.csv")
print()

# Load the CSV file
with open('csv_files/email_table.csv', 'r', newline='') as f:
    reader = csv.reader(f)
    rows = list(reader)

# Create a new Excel workbook and worksheet
wb = Workbook()
ws = wb.active

# Write the rows to the worksheet
for row in rows:
    ws.append(row)

# Remove empty rows
for i in reversed(range(1, ws.max_row + 1)):
    if all([cell.value is None for cell in ws[i]]):
        ws.delete_rows(i)

# Set the fill color and font color for the "Same Date" rows
fill = PatternFill(start_color='4f81bd', end_color='4f81bd', fill_type='solid')
font_color = 'fff2cc'

# Loop through the rows and highlight the "Same Date" rows
for row in ws.iter_rows(min_row=2, min_col=1, max_col=5):
    if row[4].value == "CONCURRENT SHOW AT SCOOT INN" or row[4].value == "CONCURRENT SHOW AT EMO'S":
        for cell in row:
            cell.fill = fill
            cell.font = cell.font.copy(color=font_color)

# Save the workbook to a file
wb.save('xlsx_files/email_table.xlsx')