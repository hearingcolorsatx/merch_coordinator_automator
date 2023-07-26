import json, logging, pyautogui, pyperclip, requests, time, webbrowser

from calendar_generator import *
from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Color
from tabulate import tabulate
from datetime import datetime, timedelta
from urllib.parse import urlparse, parse_qs

logging.basicConfig(level=logging.ERROR)

config = json.load(open("config_files/parameters.json"))
emails = json.load(open("config_files/emails.json"))
# print(emails["recipients"])

# Define the request parameters
authorize_endpoint = config["auth_endpoint"]
client_id = config["client_id"]
client_secret = config["client_secret"]
redirect_uri = config["redirect_uri"]
scope = config["scope"]
token_endpoint = config["token_endpoint"]

# Generate the authorization URL
auth_url = authorize_endpoint + "?" + "response_type=code" + "&" + "client_id=" + client_id + "&" + "redirect_uri=" + redirect_uri + "&" + "scope=" + scope

# Open the authorization URL in the default browser
webbrowser.open(auth_url)

# Wait for the user to complete the sign-in process and return to the redirect URI
redirect_uri_path = urlparse(redirect_uri).path
authorization_code = None

while authorization_code is None:
    # Wait 15s for page to load while user logs in (adjust as needed)
    time.sleep(5)
    # Press Ctrl + L to focus on URL bar (Keep browser window open on redirect page and selected!!)
    pyautogui.hotkey('ctrl', 'l')
    # Press Ctrl + C to copy URL
    pyautogui.hotkey('ctrl', 'c')
    # Get URL from clipboard
    url = pyperclip.paste()

    # Parse the code from the url string
    query_string = urlparse(url).query
    query_params = parse_qs(query_string)
    if "code" in query_params and redirect_uri_path in urlparse(url).path:
        authorization_code = query_params["code"][0]

# Exchange the authorization code for an access token
response = requests.post(token_endpoint, data={
    "grant_type": "authorization_code",
    "client_id": client_id,
    "client_secret": client_secret,
    "code": authorization_code,
    "redirect_uri": redirect_uri,
    "scope": scope
})

# Parse the response to obtain the access token
if response.status_code == 200:
    access_token = json.loads(response.text)["access_token"]
else:
    print("Error:", response.status_code, response.text)

# Set the API endpoint URL for sending email
api_url = 'https://graph.microsoft.com/v1.0/me/sendMail'

# Set the access token in the Authorization header
headers = {'Authorization': 'Bearer ' + access_token,
           'Content-Type': 'application/json'}

next_month = (datetime.now() + timedelta(days=30)).strftime("%B")

compile_info()

# Load the xlsx file
workbook = load_workbook(filename='xlsx_files/email_table.xlsx')
worksheet = workbook.active

# Read the data from the worksheet
data = []
for row in worksheet.iter_rows(min_row=1, min_col=1, max_col=5, values_only=False):
    # print(row)
    row_data = []
    for cell in row:
        value = cell.value
        font = cell.font.name
        fill = cell.fill
        color = cell.font.color.rgb
        # print(color)
        row_data.append((value, font, fill, color))
        # print(row_data)
    data.append(row_data)
# print(row_data)

# Set the email message properties
table_html = """
<table border="1" cellpadding="2" cellspacing="0">
<tr>
{header_cells}
</tr>
{data_rows}
</table>
"""
header_cells = ""
for header, font, fill, color in data[0]:
    header_cells += f"<th style='background-color:{fill.bgColor.rgb};color:{color}{header}</th>"

data_rows = ""
for row_data in data[1:]:
    row_html = ""
    for cell in row_data:
        cell_value = "" if cell[0] is None else str(cell[0])
        cell_style = f"background-color:{cell[2].bgColor.rgb};color:{cell[3]}"
        if cell_style.endswith("'>"):
            # cell_style = cell_style[:-2]
            cell_style = "background-color:00FFFFFF;color:00000000"
        # if cell[2].bgColor.rgb != "FFFFFF":
        print(cell_style)
        row_html += f"<td style='{cell_style}'>{cell_value}</td>"
        # else:
        #     print(cell[2].bgColor.rgb)
        #     row_html += f"<td>{cell_value}</td>"
    data_rows += f"<tr>{row_html}</tr>"

payload = {
    "message": {
        "subject": f"{next_month} Calendar for Emo's/Scoot Inn",
        "body": {
            "contentType": "HTML",
            "content": f'''<html>Hi All!<br/><br/>I hope everyone is well. It's that time of the month to start looking ahead to {next_month}. The following shows need to be covered:
                        <br/><br/>{table_html.format(header_cells=header_cells, data_rows=data_rows)}
                        <br/>Make note of dates where there are shows at both venues as indicated above, and let me know if there are specific shows you'd like to work. Otherwise,
                        just let me know your availability and I'll get the schedule put together in the next day or two. As always, it's first come first serve.
                        <br/><br/>The list of shows in this email is created from what is currently listed on the event calendars of the Emo's and Scoot Inn websites, so there may
                        be more shows added at a later date. I will do my best to keep you all informed of any such updates.
                        <br/><br/>If you have any questions, please let me know.
                        <br/><br/>Cheers,<br/>{emails["name"]}
                        <br/>Emo's/Scoot Inn Merch Coordinator
                        <br/>{emails["sender"]} // {emails["number"]}</html>'''
        },
        "toRecipients": emails["recipients"],
        "ccRecipients": emails["cc"]
    },
    "saveToSentItems": "true"
}

# Send the email
response = requests.post(api_url, headers=headers, data=json.dumps(payload))

# Check the response status code to see if the email was sent successfully
if response.status_code == 202:
    print("Email sent successfully!")
else:
    print("Failed to send email.", response.text)
